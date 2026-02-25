from openpyxl import load_workbook
import pandas as pd

from app.services.trade_export_service import export_to_excel_file


def test_export_to_excel_file_writes_summary_values(tmp_path):
    df = pd.DataFrame(
        [
            {
                "Price_Change_Pct": 0.1,
                "PNL_Before_Fees": 50.0,
                "Fees": -3.0,
                "PNL_Net": 47.0,
            },
            {
                "Price_Change_Pct": -0.2,
                "PNL_Before_Fees": -20.0,
                "Fees": -2.0,
                "PNL_Net": -22.0,
            },
        ]
    )
    out = tmp_path / "trades.xlsx"

    path = export_to_excel_file(df, filename=str(out))

    assert path == str(out)
    wb = load_workbook(str(out))
    ws = wb["Orders"]
    summary_row = len(df) + 2
    assert ws[f"A{summary_row}"].value == "Summary"
    assert float(ws[f"B{summary_row}"].value) == 30.0
    assert float(ws[f"C{summary_row}"].value) == -5.0
    assert float(ws[f"D{summary_row}"].value) == 25.0
    wb.close()
